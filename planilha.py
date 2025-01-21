import openpyxl
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from navegador import analisar_email, home_page, driver

#Este arquivo é responsável por ler a planilha e chamar a função analisar_email para cada email da planilha

# Caminho para o arquivo Excel
arquivo_excel = r'C:\\Analisar\\gmails.xlsx'

# Carrega a planilha com o openpyxl
workbook = openpyxl.load_workbook(arquivo_excel)
sheet = workbook.active  # Seleciona a primeira aba
def rodar():
    # Iterar pelas linhas da coluna A (emails) até o último email
    for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        gmail = linha[0].value  # Coluna A (Emails)
        linha_num = linha[0].row  # Pega o número da linha atual
        
        if gmail is None:
            print(f"Encontrou uma linha vazia na linha {linha_num}, interrompendo o loop.")
            driver.quit()
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("AutoAnalise", "Automação finalizada!")
            break  # Sai do loop se encontrar uma linha vazia

        if "@lhlaw" in gmail:
            sheet.cell(row=linha_num, column=8, value="Conta LHLAW")
            print(f"Email {gmail} é LHLAW, pulando para o próximo.")
            continue  # Pula para o próximo email se o email for LHLAW
        elif "2024" in gmail:
            sheet.cell(row=linha_num, column=8, value="Conta 2024")
            print(f"Email {gmail} é 2024, pulando para o próximo.")
            continue  # Pula para o próximo email se o email for 2024
        elif "2023" in gmail:
            sheet.cell(row=linha_num, column=8, value="Conta 2023")
            print(f"Email {gmail} é 2023, pulando para o próximo.")
            continue  # Pula para o próximo email se o email for 2023
        
        # Verificar se há um valor de email
        if gmail:
            if sheet.cell(row=linha_num, column=8).value:
                print(f"Email {gmail} já analisado, pulando para o próximo.")
                continue  # Pula para o próximo email se a célula estiver preenchida

            print(f"Processando email: {gmail}")
            resultado = analisar_email(gmail)  # Chama o navegador para analisar o email

            # Dependendo do retorno das funções, escreva na planilha
            if resultado == 'Data Reviewed':
                sheet.cell(row=linha_num, column=8, value="Já marcada como DR no GAdmin")  # Coluna H
                print('Marcado como: Data Reviewed')
            elif resultado in ['/BR/Generic Mailbox','/BR/VIALTO/Generic Mailbox', '/BR/DO_NOT_DELETE/AGTech']:
                sheet.cell(row=linha_num, column=8, value=resultado)  # Coluna H
                print('Marcado como: casos de exeção')
            elif resultado == 'Tem SD, arquivos com menos de 1 ano e google sites':
                sheet.cell(row=linha_num, column=8, value="Tem SD, arquivos com menos de 1 ano e google sites")
                print('Marcado como: Tem SD, arquivos com menos de 1 ano e google sites')
            elif resultado == 'Tem SD e arquivos a menos de 1 ano':
                sheet.cell(row=linha_num, column=8, value="Tem SD e arquivos a menos de 1 ano")
                print('Marcado como: Tem SD e arquivos a menos de 1 ano')
            elif resultado == 'Tem GSites e arquivos a menos de 1 ano':
                sheet.cell(row=linha_num, column=8, value="Tem GSites e arquivos a menos de 1 ano")
                print('Marcado como: Tem GSites e arquivos a menos de 1 ano')
            elif resultado == 'Tem SD e tem GSites':
                sheet.cell(row=linha_num, column=8, value="Tem SD e tem GSites")
                print('Marcado como: Tem SD e tem GSites')
            elif resultado == 'Tem shared drive':
                sheet.cell(row=linha_num, column=8, value="Tem shared drive")
                print('Marcado como: Tem shared drive')
            elif resultado == 'Arquivo com menos de um ano':
                sheet.cell(row=linha_num, column=8, value="Arquivo com menos de um ano")
                print('Marcado como: Arquivo com menos de um ano')
            elif resultado == 'Tem google site':
                sheet.cell(row=linha_num, column=8, value="Tem google site")
                print('Marcado como: Tem google site')
            elif resultado == 'Marcado como DR no Google':
                sheet.cell(row=linha_num, column=8, value="Marcado como DR no Google")
                sheet.cell(row=linha_num, column=6, value="DR9")
                print('Marcado como DR no Google')
            elif resultado == 'Botão Data Reviewed desabilitado':
                sheet.cell(row=linha_num, column=8, value="Botão Data Reviewed desabilitado")
                print('Botão Data Reviewed desabilitado')
            elif resultado == 'Analisar':
                sheet.cell(row=linha_num, column=8, value="Analisar Org Unit Path")
                print('Marcado como: Analisar Org Unit Path')
            else:
                sheet.cell(row=linha_num, column=8, value="Analisar")
                print('Marcado como: Analisar')

        # Salva as alterações no arquivo Excel
        workbook.save(r'C:\\Analisar\\gmails.xlsx')
        print (f'Analise do email: {gmail} feita com sucesso')
        home_page()